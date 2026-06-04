#!/usr/bin/env python3
"""Сборка Excel-отчёта из state-файла collect.py.

Листы: WB (если есть кабинет WB), Ozon (если есть), Общее — всегда.
Формат утверждён пользователем:
  - строка 1 — шапка (высота 46), строка 2 — ИТОГО, закреплены обе;
  - все Δ-колонки: любой плюс — зелёная заливка, любой минус — красная;
  - фильтр: только активные артикулы (остаток>0 ИЛИ заказы за 14 дней>0);
  - дубли с префиксами (FBS)/блок/бан объединяются в «Общее» к базовому артикулу.

Сам сверяет итоги листов с сырыми данными и печатает VERIFY OK/MISMATCH.

Использование:
  python3 build_excel.py --state <state.json> --outdir <папка>
"""
import argparse, json, re, os, sys
from datetime import date
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from common import norm, load_map, base_art
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

p = argparse.ArgumentParser()
p.add_argument("--state", required=True)
p.add_argument("--outdir", required=True)
a = p.parse_args()

S = json.load(open(a.state))
MAP = load_map(os.path.dirname(os.path.abspath(a.state)), __import__("re").sub(r"[^A-Z0-9]", "", S["brand"].upper()))
Y, P = S["yest"], S["prev"]
BRAND = S["brand"]
ddmm = lambda iso: f"{iso[8:10]}.{iso[5:7]}"
Y_L, P_L = ddmm(Y), ddmm(P)

def wb_active(nm):
    nm = str(nm)
    return S.get("wb_stocks", {}).get(nm, 0) > 0 or S.get("wb_orders14", {}).get(nm, 0) > 0

def oz_active(offer, sku):
    return (S.get("oz_stocks", {}) or {}).get(offer, 0) > 0 or \
           (S.get("oz_orders14", {}) or {}).get(str(sku), 0) > 0

wb_rows = []
if S.get("has_wb"):
    for c in S["wb_cards"]:
        if not wb_active(c["nmID"]): continue
        f = S["wb_funnel"].get(str(c["nmID"]), {})
        y, pp = f.get(Y, {}), f.get(P, {})
        if not y and not pp: continue
        wb_rows.append({"art": c["vendorCode"], "nm": c["nmID"],
            "o_y": y.get("orders", 0), "o_p": pp.get("orders", 0),
            "v_y": y.get("open", 0),   "v_p": pp.get("open", 0),
            "c_y": y.get("cart", 0),   "c_p": pp.get("cart", 0)})
    wb_rows.sort(key=lambda r: -r["o_y"])

oz_rows = []
if S.get("has_oz"):
    for sku, v in (S.get("ozon") or {}).items():
        if not oz_active(v.get("offer_id", ""), sku): continue
        y, pp = v.get(Y, {}), v.get(P, {})
        if not y and not pp: continue
        oz_rows.append({"art": v.get("offer_id") or v["name"], "sku": sku,
            "o_y": y.get("orders", 0),    "o_p": pp.get("orders", 0),
            "sh_y": y.get("view", 0),     "sh_p": pp.get("view", 0),
            "cl_y": y.get("view_pdp", 0), "cl_p": pp.get("view_pdp", 0),
            "c_y": y.get("tocart", 0),    "c_p": pp.get("tocart", 0)})
    oz_rows.sort(key=lambda r: -r["o_y"])

HDR_FILL = PatternFill("solid", start_color="1F4E78")
HDR_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BASE_FONT = Font(name="Arial", size=10)
BOLD = Font(name="Arial", bold=True, size=10)
TOTAL_FILL = PatternFill("solid", start_color="DDEBF7")
THIN = Border(*[Side(style="thin", color="B0B0B0")]*4)
GOOD_FILL = PatternFill("solid", start_color="C6EFCE")
BAD_FILL = PatternFill("solid", start_color="FFC7CE")
GOOD_FONT = Font(name="Arial", size=10, color="006100")
BAD_FONT = Font(name="Arial", size=10, color="9C0006")

def sheet_header(ws, cols, widths):
    for i, (t, w) in enumerate(zip(cols, widths), 1):
        c = ws.cell(row=1, column=i, value=t)
        c.font = HDR_FONT; c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"
    ws.row_dimensions[1].height = 46

def delta_fmt(c): c.number_format = '+#,##0;-#,##0;"—"'
def pct_fmt(c):   c.number_format = '+0.0%;-0.0%;"—"'

def cf_red_green(ws, col, first, last):
    rng = f"{col}{first}:{col}{last}"
    ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThan", formula=["0"],
                                                  fill=GOOD_FILL, font=GOOD_FONT))
    ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["0"],
                                                  fill=BAD_FILL, font=BAD_FONT))

def total_style(ws, ncols):
    for i in range(1, ncols+1):
        c = ws.cell(row=2, column=i); c.font = BOLD; c.fill = TOTAL_FILL; c.border = THIN

book = Workbook()
first_sheet = True

if wb_rows:
    ws = book.active if first_sheet else book.create_sheet()
    ws.title = "WB"; first_sheet = False
    cols = ["Артикул", "nmID", f"Заказы вчера\n({Y_L})", f"Заказы позавчера\n({P_L})", "Δ заказы", "Δ %",
            "Переходы в карточку\nвчера", "Переходы\nпозавчера", "Δ переходы",
            "Корзины вчера", "Корзины позавчера", "Δ корзины"]
    sheet_header(ws, cols, [42, 11, 12, 13, 10, 9, 13, 12, 11, 12, 13, 11])
    first, last = 3, 2 + len(wb_rows)
    r = 3
    for row in wb_rows:
        vals = [row["art"], row["nm"], row["o_y"], row["o_p"], f"=C{r}-D{r}",
                f"=IF(D{r}=0,\"\",(C{r}-D{r})/D{r})", row["v_y"], row["v_p"], f"=G{r}-H{r}",
                row["c_y"], row["c_p"], f"=J{r}-K{r}"]
        for i, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=i, value=v); c.font = BASE_FONT; c.border = THIN
        delta_fmt(ws.cell(row=r, column=5)); pct_fmt(ws.cell(row=r, column=6))
        delta_fmt(ws.cell(row=r, column=9)); delta_fmt(ws.cell(row=r, column=12))
        r += 1
    ws.cell(row=2, column=1, value="ИТОГО WB")
    for col, L in [(3,"C"),(4,"D"),(5,"E"),(7,"G"),(8,"H"),(9,"I"),(10,"J"),(11,"K"),(12,"L")]:
        ws.cell(row=2, column=col, value=f"=SUM({L}{first}:{L}{last})")
    ws.cell(row=2, column=6, value="=IF(D2=0,\"\",(C2-D2)/D2)")
    delta_fmt(ws.cell(row=2, column=5)); pct_fmt(ws.cell(row=2, column=6))
    delta_fmt(ws.cell(row=2, column=9)); delta_fmt(ws.cell(row=2, column=12))
    total_style(ws, 12)
    for L in ("E", "F", "I", "L"):
        cf_red_green(ws, L, 2, last)

if oz_rows:
    ws = book.active if first_sheet else book.create_sheet()
    ws.title = "Ozon"; first_sheet = False
    cols = ["Артикул", "SKU Ozon", f"Заказы вчера\n({Y_L})", f"Заказы позавчера\n({P_L})", "Δ заказы", "Δ %",
            "Показы вчера", "Показы позавчера", "Δ показы",
            "Просмотры карточки\nвчера", "Просмотры\nпозавчера", "Δ просмотры",
            "Корзины вчера", "Корзины позавчера", "Δ корзины"]
    sheet_header(ws, cols, [42, 13, 12, 13, 10, 9, 13, 14, 12, 14, 13, 12, 12, 13, 11])
    first, last = 3, 2 + len(oz_rows)
    r = 3
    for row in oz_rows:
        vals = [row["art"], int(row["sku"]), row["o_y"], row["o_p"], f"=C{r}-D{r}",
                f"=IF(D{r}=0,\"\",(C{r}-D{r})/D{r})", row["sh_y"], row["sh_p"], f"=G{r}-H{r}",
                row["cl_y"], row["cl_p"], f"=J{r}-K{r}", row["c_y"], row["c_p"], f"=M{r}-N{r}"]
        for i, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=i, value=v); c.font = BASE_FONT; c.border = THIN
        delta_fmt(ws.cell(row=r, column=5)); pct_fmt(ws.cell(row=r, column=6))
        delta_fmt(ws.cell(row=r, column=9)); delta_fmt(ws.cell(row=r, column=12)); delta_fmt(ws.cell(row=r, column=15))
        r += 1
    ws.cell(row=2, column=1, value="ИТОГО Ozon")
    for col, L in [(3,"C"),(4,"D"),(5,"E"),(7,"G"),(8,"H"),(9,"I"),(10,"J"),(11,"K"),(12,"L"),(13,"M"),(14,"N"),(15,"O")]:
        ws.cell(row=2, column=col, value=f"=SUM({L}{first}:{L}{last})")
    ws.cell(row=2, column=6, value="=IF(D2=0,\"\",(C2-D2)/D2)")
    delta_fmt(ws.cell(row=2, column=5)); pct_fmt(ws.cell(row=2, column=6))
    delta_fmt(ws.cell(row=2, column=9)); delta_fmt(ws.cell(row=2, column=12)); delta_fmt(ws.cell(row=2, column=15))
    total_style(ws, 15)
    for L in ("E", "F", "I", "L", "O"):
        cf_red_green(ws, L, 2, last)

# --- Общее ---
agg = {}
for row in wb_rows:
    g = agg.setdefault(base_art(row["art"], MAP), {"y": 0, "p": 0})
    g["y"] += row["o_y"]; g["p"] += row["o_p"]
for row in oz_rows:
    g = agg.setdefault(base_art(row["art"], MAP), {"y": 0, "p": 0})
    g["y"] += row["o_y"]; g["p"] += row["o_p"]
items = sorted(agg.items(), key=lambda kv: -kv[1]["y"])

ws = book.active if first_sheet else book.create_sheet()
ws.title = "Общее"
cols = ["Артикул (база)", f"Заказы ВСЕГО\nвчера ({Y_L})", f"Заказы ВСЕГО\nпозавчера ({P_L})", "Δ всего", "Δ %"]
sheet_header(ws, cols, [44, 14, 15, 10, 9])
first, last = 3, 2 + len(items)
r = 3
for k, g in items:
    vals = [k, g["y"], g["p"], f"=B{r}-C{r}", f"=IF(C{r}=0,\"\",(B{r}-C{r})/C{r})"]
    for i, v in enumerate(vals, 1):
        c = ws.cell(row=r, column=i, value=v); c.font = BASE_FONT; c.border = THIN
    delta_fmt(ws.cell(row=r, column=4)); pct_fmt(ws.cell(row=r, column=5))
    r += 1
ws.cell(row=2, column=1, value="ИТОГО")
for col, L in [(2,"B"),(3,"C"),(4,"D")]:
    ws.cell(row=2, column=col, value=f"=SUM({L}{first}:{L}{last})")
ws.cell(row=2, column=5, value="=IF(C2=0,\"\",(B2-C2)/C2)")
delta_fmt(ws.cell(row=2, column=4)); pct_fmt(ws.cell(row=2, column=5))
total_style(ws, 5)
for L in ("D", "E"):
    cf_red_green(ws, L, 2, last)

src = []
if wb_rows: src.append("WB — Воронка продаж (Seller Analytics API)")
if oz_rows: src.append("Ozon — Seller API analytics")
note = ws.cell(row=last+2, column=1, value=("Источники: " + "; ".join(src) + ". "
    "Только активные артикулы (остаток > 0 или заказы за 14 дней > 0). "
    "Дубли с префиксами (FBS)/блок/бан объединены к базовому артикулу."))
note.font = Font(name="Arial", size=9, italic=True, color="808080")

# «Общее» — всегда первый лист
book.move_sheet("Общее", offset=-(len(book.sheetnames) - 1))
book.active = 0
os.makedirs(a.outdir, exist_ok=True)
out = os.path.join(a.outdir, f"Отчёт_{BRAND}_{Y_L}.xlsx")
book.save(out)

# --- самопроверка: фильтр не должен терять заказы активных артикулов ---
wb_lost = [(nm, d.get(Y, {}).get("orders", 0)) for nm, d in (S.get("wb_funnel") or {}).items()
           if d.get(Y, {}).get("orders", 0) > 0 and not wb_active(nm)]
oz_lost = [(sk, v.get(Y, {}).get("orders", 0)) for sk, v in (S.get("ozon") or {}).items()
           if v.get(Y, {}).get("orders", 0) > 0 and not oz_active(v.get("offer_id", ""), sk)]
sum_wb = sum(r["o_y"] for r in wb_rows); sum_oz = sum(r["o_y"] for r in oz_rows)
sum_all = sum(g["y"] for _, g in items)
verify = (sum_all == sum_wb + sum_oz) and not wb_lost and not oz_lost
print(f"saved {out}")
print(f"WB строк: {len(wb_rows)} | Ozon строк: {len(oz_rows)} | заказы вчера: WB {sum_wb}, Ozon {sum_oz}, всего {sum_all}")
if wb_lost or oz_lost:
    print(f"ВНИМАНИЕ, потеряны заказы фильтром: WB {wb_lost} Ozon {oz_lost}")
print("VERIFY", "OK" if verify else "MISMATCH")
